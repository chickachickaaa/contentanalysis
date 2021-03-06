#!/usr/bin/env python3

"""
cat_newfile_upgrades.py is used to create four charts
from the url data in the csv files created from the internal tool 
called datashot (ie. grapeshot's url categorization tool). 
Each row in the csv file represents a URL with it's corresponding grapeshot
URL categorization data. We are working to incoropate the latest
feedback from Sales here, cat.py is perfect pre-sales. **It has the notes to create
tests from.**
Author: Taylor Higgins
Last modified: 2/28/2020
"""
import sys
import re
import csv
import pandas as pd
import openpyxl
import operator
import collections
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint

#should I up top here determine if it's a https or www so we can have a value to decide which logic to use going forward? 
#instead of asking in each function? 

def url_path_parser():
    """Loop to parse all the URL path sections from the existing URLs. This will be used for the third chart which I haven't started yet. """
    #Todo: Remove intermediary tuple section. That caused an error when appending so I'm going to leave it for now. 
    
    #http_true = 'unknown'

    for row in cat_sheet.iter_rows(min_col=1, min_row=2, max_col=1, max_row=cat_sheet.max_row):    
        for cell in row:
            #if column1 url in sheet 1
            #contains http://
            #do this logic for http:// url's
            #this below works for https
            #split_cell_list = cell.value.split('/')[3:]
            #else:
            #do this logic for www url's
            #print(type(cell.value)) #string
            #this below works for www.
            if 'http' in cell.value:
                split_cell_list = cell.value.split('/')[3:]
                http_true = 1
            else:
                split_cell_list = cell.value.split('/')[1:]
                http_true = 0
            tuple_section = (split_cell_list,)
            #print(http_true)
            #print(type(tuple_section)) #tuple
            for row in tuple_section:
                #print(type(row)) #list
                #print(row)
                sections_sheet.append(row)
                #data_sheet.append(row)
    return(http_true)

def url_totals():

    """Find Totals for the charts. Categorized URLs total comes from the total row count in the categorized sheet 
    or by going through all the non gv_ and gx_ segments in the all sheet. Unsafe URLs Total
    comes from any url starting with gv_. """             

    unsafe_total = 0
    null_total = 0
    safe_segment_dict = {}
    section_dict = {}

    #Starting loop to find uniquely null URL rows.
    for row in cat_sheet.iter_rows(min_col=2, min_row=2, max_col=cat_sheet.max_column, max_row=cat_sheet.max_row): 
        for cell in row:
            unsafe_segment = re.compile(r'^gx_')
            try:
                check_null = unsafe_segment.search(cell.value)
                if check_null is not None: 
                    null_total +=1 
                    #We remove break because we want any gx to be returned. 
                else:
                    continue

            except TypeError:
                #This should pass from the blank cell to the next in the row, then finally to next URL row.
                pass 

    #Starting loop to find uniquely unsafe URL rows.
    for row in cat_sheet.iter_rows(min_col=2, min_row=2, max_col=cat_sheet.max_column, max_row=cat_sheet.max_row): 
        for cell in row:
            #write this out so we ignore gv_safe
            fake_unsafe_segment = re.compile(r'^gv_safe')
            unsafe_segment = re.compile(r'^gv_')
            try:
                check_fake = fake_unsafe_segment.search(cell.value)
                check_safety = unsafe_segment.search(cell.value)
                if check_fake is not None:
                    #print(cell.value)
                    break
                elif check_safety is not None: 
                    #print(cell.value)
                    unsafe_total +=1 
                    #We put a break here so that we don't over count URLs that were categorized multiple times as unsafe.
                    break 

            except TypeError:
                #This should pass from the blank cell to the next in the row, then finally to next URL row.
                pass 

    #Starting loop to create a dict of absolute count of segment appearances regardless of unique URL row. 
    for row in cat_sheet.iter_rows(min_col=2, min_row=2, max_col=cat_sheet.max_column, max_row=cat_sheet.max_row):
        for cell in row:
            safe_segment = re.compile(r'^gs_')
            try:
                check_safety = safe_segment.search(cell.value)
                if check_safety is not None:
                    new_segment = cell.value
                    if new_segment in safe_segment_dict: 
                        #If the segment is in the dict already add one to the current value.
                        safe_segment_dict[new_segment] = safe_segment_dict[new_segment] + 1
                    else:
                        #If the segment is not in the dict already then add it and make the value 1.
                        safe_segment_dict[new_segment] = 1
            except TypeError:
                pass

    #Here we are sorting the dict of safe segments in descending order by the value of the dict.   
    sorted_safe_seg_dict = dict(sorted(safe_segment_dict.items(), key=operator.itemgetter(1), reverse=True))
    #print(sorted_safe_seg_dict)
    #Here we are turning the sorted dict into a list so we can later append it by row and create the segment bar chart.
    safe_segment_list = list(sorted_safe_seg_dict.items())
    #print(safe_segment_list[0][0])

    #Starting loop to create a dict of absolute count of section appearances regardless of unique URL row or safety.
    for row in sections_sheet.iter_rows(min_col=1, min_row=2, max_col=sections_sheet.max_column, max_row=sections_sheet.max_row):
        #print(row)
        for cell in row:
            #print(cell.value)
            try:
                if cell.value in section_dict:
                    #If section is in the dict already add one to the current value. 
                    section_dict[cell.value] = section_dict[cell.value] + 1
                    #print(cell.value)
                else:
                    #If the segment is not in the dict already then add it and make the value 1.
                    section_dict[cell.value] = 1
            except TypeError:
                pass

    #Here we are sorting the dict of sections in descending order by the value of the dict.
    sorted_section_dict = dict(sorted(section_dict.items(), key=operator.itemgetter(1), reverse=True))

    #print(sorted_section_dict)

    #Here we are turning the sorted dict into a list so we can later append it by row to the file. 
    section_list = list(sorted_section_dict.items())
    

    
    #Calculate total categorized URLs. 
    row_count = cat_sheet.max_row 
    total_cat_urls = row_count - 1 - null_total
    safe_total = total_cat_urls - unsafe_total

    charts_sheet['A1'] = 'Total URLs Categorized' 
    charts_sheet['B1'] = total_cat_urls
    # print("null total is {}".format(null_total))
    # print("unsafe total is {}".format(unsafe_total))
    # print("safe total is {}".format(safe_total))

    return(null_total, unsafe_total, safe_total, safe_segment_list, section_list)

def section_many_segments(section_list):
    """Create the Incremental Inventory Bar Chart"""
    #let's print this to a sheet to look into it section list
    #Here we pick the most popular section from the section list. 
    #Note we might need try except for cover for errors
    #print(section_list)
    for row in section_list:
        section_list_sheet.append(row)

    print(section_list[0][0])
    most_popular_section = section_list[1][0] #do the top 3
    print(most_popular_section)
    second_most_popular_section = section_list[2][0]
    print(second_most_popular_section)
    third_most_popular_section = section_list[3][0]
    print(third_most_popular_section)
    print(section_list[4][0])
    #print(section_list[2][0])
    hardchart1a_sheet['B1'] = most_popular_section
    hardchart1b_sheet['B1'] = second_most_popular_section
    hardchart1c_sheet['B1'] = third_most_popular_section

    popular_segment_count = {}
    second_most_popular_segment_count = {}
    third_most_popular_segment_count = {}

    #Here we loop through all rows in column 1 or URLs in cat_sheet
    #Currently this will loop through everything so there could be an issue.
    #If cell.value in a non column 1 col contains most_popular_section
    #for chart 1 option 1
    for row in cat_sheet.iter_rows(min_row=2,min_col=1,max_col=cat_sheet.max_column, max_row=cat_sheet.max_row):
        #For any cell in column 1 that contains the most popular section
        for cell in row:
            try:
                if most_popular_section in cell.value:
                    #for every 3 column after column 2 add the value of the cell to a dict and start counting it 
                    column_list = [2,5,8,11,14,17,20,23,26,29,32,35,38,41,44]
                    #dynamically loop through all segment columns
                    for column_number in column_list:
                        popular_segment = cat_sheet.cell(row=cell.row, column=column_number).value
                        #go through all cells in that row and pull out all gs segments
                        safe_segment = re.compile(r'^gs_') 
                        #try adding in the two _ thing like safe_child_segment = r'^gs_wildcard_'
                        try:
                            #check if that segment is in the dict yet 
                            check_seg_safety = safe_segment.search(popular_segment)
                            if check_seg_safety is not None:
                                if popular_segment in popular_segment_count:
                                    popular_segment_count[popular_segment] = popular_segment_count[popular_segment] + 1
                                else:
                                    popular_segment_count[popular_segment] = 1
                        except TypeError:
                            pass
                else:
                    continue
            except TypeError:
                pass

    popular_section_key = [most_popular_section]

    #this goes into a nested dict. {1_section:{1seg:1segcount,2seg:2segcount,3seg:3segcount}}
    nested_popular_segment_count = dict.fromkeys(popular_section_key,popular_segment_count)

    #then we can go through similarly as we did in url_totals of sorting
    sorted_popular_segment_count = dict(sorted(popular_segment_count.items(), key=operator.itemgetter(1), reverse=True))
    #making a list popular_segment_count
    sorted_popular_seg_list = list(sorted_popular_segment_count.items())


    #appending sorted_popular_seg_list to sections sheet
    #for chart 1 option 1
    for row in sorted_popular_seg_list:
        #sections_sheet.append(row)
        #data_sheet.append(row)
        hardchart1a_sheet.append(row)

#for chart 1 option 2
    for row in cat_sheet.iter_rows(min_row=2,min_col=1,max_col=cat_sheet.max_column, max_row=cat_sheet.max_row):
        #For any cell in column 1 that contains the most popular section
        for cell in row:
            try:
                if second_most_popular_section in cell.value:
                    #for every 3 column after column 2 add the value of the cell to a dict and start counting it 
                    column_list = [2,5,8,11,14,17,20,23,26,29,32,35,38,41,44]
                    #dynamically loop through all segment columns
                    for column_number in column_list:
                        popular_segment = cat_sheet.cell(row=cell.row, column=column_number).value
                        #go through all cells in that row and pull out all gs segments
                        safe_segment = re.compile(r'^gs_')
                        try:
                            #check if that segment is in the dict yet 
                            check_seg_safety = safe_segment.search(popular_segment)
                            if check_seg_safety is not None:
                                if popular_segment in second_most_popular_segment_count:
                                    second_most_popular_segment_count[popular_segment] = second_most_popular_segment_count[popular_segment] + 1
                                else:
                                    second_most_popular_segment_count[popular_segment] = 1
                        except TypeError:
                            pass
                else:
                    continue
            except TypeError:
                pass

    second_most_popular_section_key = [second_most_popular_section]

    #this goes into a nested dict. {1_section:{1seg:1segcount,2seg:2segcount,3seg:3segcount}}
    nested_second_popular_segment_count = dict.fromkeys(second_most_popular_section_key,second_most_popular_segment_count)

    #then we can go through similarly as we did in url_totals of sorting
    sorted_second_popular_segment_count = dict(sorted(second_most_popular_segment_count.items(), key=operator.itemgetter(1), reverse=True))
    #making a list popular_segment_count
    sorted_second_popular_seg_list = list(sorted_second_popular_segment_count.items())


    #for chart 1 option 2
    for row in sorted_second_popular_seg_list:
        hardchart1b_sheet.append(row)

#for chart 1 option 3
    for row in cat_sheet.iter_rows(min_row=2,min_col=1,max_col=cat_sheet.max_column, max_row=cat_sheet.max_row):
        #For any cell in column 1 that contains the most popular section
        for cell in row:
            try:
                if third_most_popular_section in cell.value:
                    #for every 3 column after column 2 add the value of the cell to a dict and start counting it 
                    column_list = [2,5,8,11,14,17,20,23,26,29,32,35,38,41,44]
                    #dynamically loop through all segment columns
                    for column_number in column_list:
                        popular_segment = cat_sheet.cell(row=cell.row, column=column_number).value
                        #go through all cells in that row and pull out all gs segments
                        safe_segment = re.compile(r'^gs_')
                        try:
                            #check if that segment is in the dict yet 
                            check_seg_safety = safe_segment.search(popular_segment)
                            if check_seg_safety is not None:
                                if popular_segment in third_most_popular_segment_count:
                                    third_most_popular_segment_count[popular_segment] = third_most_popular_segment_count[popular_segment] + 1
                                else:
                                    third_most_popular_segment_count[popular_segment] = 1
                        except TypeError:
                            pass
                else:
                    continue
            except TypeError:
                pass

    third_most_popular_section_key = [third_most_popular_section]

    #this goes into a nested dict. {1_section:{1seg:1segcount,2seg:2segcount,3seg:3segcount}}
    nested_third_popular_segment_count = dict.fromkeys(third_most_popular_section_key,third_most_popular_segment_count)

    #then we can go through similarly as we did in url_totals of sorting
    sorted_third_popular_segment_count = dict(sorted(third_most_popular_segment_count.items(), key=operator.itemgetter(1), reverse=True))
    #making a list popular_segment_count
    sorted_third_popular_seg_list = list(sorted_third_popular_segment_count.items())



    #for chart 1 option 3
    for row in sorted_third_popular_seg_list:
        hardchart1c_sheet.append(row)

    return(most_popular_section, second_most_popular_section, third_most_popular_section)

def segment_many_sections(safe_segment_list, http_true):
    """Create SubDomain Bar Chart"""
    #print(safe_segment_list)
    #chart 2 option 1
    most_popular_segment = safe_segment_list[0][0]
    # print(safe_segment_list[0][0])
    # print(safe_segment_list[1][0])
    # print(safe_segment_list[2][0])
    hardchart2a_sheet['B1'] = most_popular_segment
    popular_section_count = {}

    #Here we loop through all rows/segment columns in cat sheets to find any rows with the popular segment 
    for row in cat_sheet.iter_rows(min_row=2,min_col=1,max_col=cat_sheet.max_column, max_row=cat_sheet.max_row):
        for cell in row:
            try:
                if most_popular_segment in cell.value:
                #When we find a row with the popular segment
                #We need to find what section it is in
                #by going to first column of that row
                #or row number
                #finding section info from parser or url and parse again
                    #print(row[0].value)
                    url_to_be_parsed_again = row[0].value
                    #print(url_to_be_parsed_again)
                    #if column1 url in sheet 1
                    #contains http://
                    #do this logic for http:// url's
                    if http_true == 1:
                        popular_section_entire = url_to_be_parsed_again.split('/')[3:]
                        popular_section = popular_section_entire[0]
                        #print("hi we're in the if")

                        if popular_section in popular_section_count:
                            popular_section_count[popular_section] = popular_section_count[popular_section] + 1
                            #print("hi we're in the nested if")
                        else:
                            popular_section_count[popular_section] = 1
                            #print("hi we're in the nested else")
                    #else:
                    #do this logic for www url's
                    else:
                        popular_section_entire = url_to_be_parsed_again.split('/')
                        popular_section = popular_section_entire[1]
                        #print("hi we're in the else")

                        if popular_section in popular_section_count:
                            popular_section_count[popular_section] = popular_section_count[popular_section] + 1
                            #print("hi we're in the nested if")
                        else:
                            popular_section_count[popular_section] = 1
                            #print("hi we're in the nested else")
                else:       
                    continue
            except TypeError:
                pass
    #print(most_popular_segment)
    #print(popular_section_count)

    #then we can go through similarly as we did in the first chart 
    sorted_popular_section_count = dict(sorted(popular_section_count.items(), key=operator.itemgetter(1), reverse=True))
    #making a list popular_section_count
    sorted_popular_section_count = list(sorted_popular_section_count.items())
    #print(type(sorted_popular_section_count))
    #print((sorted_popular_section_count))
    for row in sorted_popular_section_count:
        hardchart2a_sheet.append(row)

    #chart 2 option 2
    second_most_popular_segment = safe_segment_list[1][0]
    # print(safe_segment_list[0][0])
    # print(safe_segment_list[1][0])
    # print(safe_segment_list[2][0])
    hardchart2b_sheet['B1'] = second_most_popular_segment
    popular_section_count = {}

    #Here we loop through all rows/segment columns in cat sheets to find any rows with the popular segment 
    for row in cat_sheet.iter_rows(min_row=2,min_col=1,max_col=cat_sheet.max_column, max_row=cat_sheet.max_row):
        for cell in row:
            try:
                if second_most_popular_segment in cell.value:
                #When we find a row with the popular segment
                #We need to find what section it is in
                #by going to first column of that row
                #or row number
                #finding section info from parser or url and parse again
                    #print(row[0].value)
                    url_to_be_parsed_again = row[0].value
                    #print(url_to_be_parsed_again)
                    #if column1 url in sheet 1
                    #contains http://
                    #do this logic for http:// url's
                    if http_true == 1:
                        popular_section_entire = url_to_be_parsed_again.split('/')[3:]
                        popular_section = popular_section_entire[0]
                        #print("hi we're in the if")

                        if popular_section in popular_section_count:
                            popular_section_count[popular_section] = popular_section_count[popular_section] + 1
                            #print("hi we're in the nested if")
                        else:
                            popular_section_count[popular_section] = 1
                            #print("hi we're in the nested else")
                    #else:
                    #do this logic for www url's
                    else:
                        popular_section_entire = url_to_be_parsed_again.split('/')
                        popular_section = popular_section_entire[1]
                        #print("hi we're in the else")

                        if popular_section in popular_section_count:
                            popular_section_count[popular_section] = popular_section_count[popular_section] + 1
                            #print("hi we're in the nested if")
                        else:
                            popular_section_count[popular_section] = 1
                            #print("hi we're in the nested else")
                else:       
                    continue
            except TypeError:
                pass
    #print(most_popular_segment)
    #print(popular_section_count)

    #then we can go through similarly as we did in the first chart 
    sorted_popular_section_count = dict(sorted(popular_section_count.items(), key=operator.itemgetter(1), reverse=True))
    #making a list popular_section_count
    sorted_popular_section_count = list(sorted_popular_section_count.items())
    #print(type(sorted_popular_section_count))
    #print((sorted_popular_section_count))
    for row in sorted_popular_section_count:
        hardchart2b_sheet.append(row)

    #chart 2 option 3
    third_most_popular_segment = safe_segment_list[2][0]
    # print(safe_segment_list[0][0])
    # print(safe_segment_list[1][0])
    # print(safe_segment_list[2][0])
    hardchart2c_sheet['B1'] = third_most_popular_segment
    popular_section_count = {}

    #Here we loop through all rows/segment columns in cat sheets to find any rows with the popular segment 
    for row in cat_sheet.iter_rows(min_row=2,min_col=1,max_col=cat_sheet.max_column, max_row=cat_sheet.max_row):
        for cell in row:
            try:
                if third_most_popular_segment in cell.value:
                #When we find a row with the popular segment
                #We need to find what section it is in
                #by going to first column of that row
                #or row number
                #finding section info from parser or url and parse again
                    #print(row[0].value)
                    url_to_be_parsed_again = row[0].value
                    #print(url_to_be_parsed_again)
                    #if column1 url in sheet 1
                    #contains http://
                    #do this logic for http:// url's
                    if http_true == 1:
                        popular_section_entire = url_to_be_parsed_again.split('/')[3:]
                        popular_section = popular_section_entire[0]
                        #print("hi we're in the if")

                        if popular_section in popular_section_count:
                            popular_section_count[popular_section] = popular_section_count[popular_section] + 1
                            #print("hi we're in the nested if")
                        else:
                            popular_section_count[popular_section] = 1
                            #print("hi we're in the nested else")
                    #else:
                    #do this logic for www url's
                    else:
                        popular_section_entire = url_to_be_parsed_again.split('/')
                        popular_section = popular_section_entire[1]
                        #print("hi we're in the else")

                        if popular_section in popular_section_count:
                            popular_section_count[popular_section] = popular_section_count[popular_section] + 1
                            #print("hi we're in the nested if")
                        else:
                            popular_section_count[popular_section] = 1
                            #print("hi we're in the nested else")
                else:       
                    continue
            except TypeError:
                pass
    #print(most_popular_segment)
    #print(popular_section_count)

    #then we can go through similarly as we did in the first chart 
    sorted_popular_section_count = dict(sorted(popular_section_count.items(), key=operator.itemgetter(1), reverse=True))
    #making a list popular_section_count
    sorted_popular_section_count = list(sorted_popular_section_count.items())
    #print(type(sorted_popular_section_count))
    #print((sorted_popular_section_count))
    for row in sorted_popular_section_count:
        hardchart2c_sheet.append(row)

    return(most_popular_segment, second_most_popular_segment, third_most_popular_segment)

def section_bar_chart(most_popular_section, second_most_popular_section, third_most_popular_section ):

    #here is a bar chart that I think is best and simplest
    #bar chart 1 option 1
    bar = BarChart()
    labels = Reference(hardchart1a_sheet, min_col=1, min_row=4, max_row=14)
    data = Reference(worksheet=hardchart1a_sheet, min_col=2, min_row=4, max_row=14)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = 'the "{}" section has many segments'.format(most_popular_section) 
    hardchart1a_sheet.add_chart(bar, "E3")

    #bar chart 1 option 2
    bar = BarChart()
    labels = Reference(hardchart1b_sheet, min_col=1, min_row=4, max_row=14)
    data = Reference(worksheet=hardchart1b_sheet, min_col=2, min_row=4, max_row=14)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = 'the "{}" section has many segments'.format(second_most_popular_section) 
    hardchart1b_sheet.add_chart(bar, "E3")

    #bar chart 1 option 3
    bar = BarChart()
    labels = Reference(hardchart1c_sheet, min_col=1, min_row=4, max_row=14)
    data = Reference(worksheet=hardchart1c_sheet, min_col=2, min_row=4, max_row=14)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = 'the "{}" section has many segments'.format(third_most_popular_section) 
    hardchart1c_sheet.add_chart(bar, "E3")

def segment_bar_chart(most_popular_segment,second_most_popular_segment, third_most_popular_segment):

    #here is a bar chart for chart 2
    bar = BarChart()
    labels = Reference(hardchart2a_sheet, min_col=1, min_row=4, max_row=14)
    data = Reference(worksheet=hardchart2a_sheet, min_col=2, min_row=4, max_row=14)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = 'sections where the "{}" segment is'.format(most_popular_segment) 
    hardchart2a_sheet.add_chart(bar, "E3")

    #here is a bar chart for chart 2
    bar = BarChart()
    labels = Reference(hardchart2b_sheet, min_col=1, min_row=4, max_row=14)
    data = Reference(worksheet=hardchart2b_sheet, min_col=2, min_row=4, max_row=14)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = 'sections where the "{}" segment is'.format(second_most_popular_segment) 
    hardchart2b_sheet.add_chart(bar, "E3")

    #here is a bar chart for chart 2
    bar = BarChart()
    labels = Reference(hardchart2c_sheet, min_col=1, min_row=4, max_row=14)
    data = Reference(worksheet=hardchart2c_sheet, min_col=2, min_row=4, max_row=14)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = 'sections where the "{}" segment is'.format(third_most_popular_segment) 
    hardchart2c_sheet.add_chart(bar, "E3")



def safe_pie_chart(null_total, unsafe_total, safe_total):
    """Create Safe Unsafe Pie Chart"""

    #Here we create the data table and appending it in the Charts Sheet.
    #Todo: since only two items just call append twice
    #TODO: make this a percent! 
    safe_unsafe_data = [['Safe', safe_total], ['Unsafe', unsafe_total]]
    for row in safe_unsafe_data:
        charts_sheet.append(row)

    #Here we set up the pie chart using openpyxl. 
    pie = PieChart()
    labels = Reference(charts_sheet, min_col=1, min_row=2, max_row=3)
    data = Reference(charts_sheet, min_col=2, min_row=1, max_row=3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Unsafe URLs"
    charts_sheet.add_chart(pie, "H2")
    #all_charts_sheet.add_chart(pie, "H10")

def popular_bar_chart(safe_segment_list):
    """Create Popular Segments Bar Chart"""

    charts_sheet['A5'] = 'Segment'
    charts_sheet['B5'] = 'Total Appearances'

    #Here we append the total segment data to the Charts Sheet.
    for row in safe_segment_list:
        charts_sheet.append(row)

    #Here we set up a bar chart using openpyxl.
    bar = BarChart()
    labels = Reference(charts_sheet, min_col=1, min_row=6, max_row=21)
    data = Reference(worksheet=charts_sheet, min_col=2, min_row=6, max_row=21)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = 'Total Appearances'
    charts_sheet.add_chart(bar, "H20")
    #all_charts_sheet.add_chart(bar, "H20" )


if __name__ == '__main__': 

    print('The context script is working, one moment please...')
    #Open csv file from Datashot using the command line argument.
    name = sys.argv[1] #meredith_cat_500.txt.cat.csv
    cleaned_name = name.split(".")
    shorter_name = cleaned_name[0] #TEST THIS
    #Transform the csv from datashot to xlsx
    #issue with unshortened so added sep and header, could maybe add: error_bad_lines=False
    #added engine='python' from error suggestion:  ParserWarning: Falling back to the 'python' engine because the 'c' engine does not support regex separator
    read_file = pd.read_csv(name, delimiter=',', header=None, engine='python', names=list(range(500))) 
    read_file.to_excel("{}{}{}".format("charts_",shorter_name,".xlsx"),index=None, header=False)
    filename = "{}{}{}".format("charts_",shorter_name,".xlsx")
    workbook = load_workbook(filename) 

    #Add sheets for data, URL sections and charts.
    #we might make more tabs for certain data 
    #workbook.create_sheet('Chart Data')
    workbook.create_sheet('URL Sections')
    #creating this to solve a bug can remove later
    workbook.create_sheet('Section List')
    workbook.create_sheet('Chart 1a')
    workbook.create_sheet('Chart 1b')
    workbook.create_sheet('Chart 1c')
    workbook.create_sheet('Chart 2a')
    workbook.create_sheet('Chart 2b')
    workbook.create_sheet('Chart 2c')
    workbook.create_sheet('Chart 3 and 4')

    #Create new sheets for sections and charts and make variables of each sheet we'll need. 
    cat_sheet = workbook['Sheet1']
    #data_sheet = workbook['Chart Data']
    sections_sheet = workbook['URL Sections']
    section_list_sheet = workbook['Section List']
    hardchart1a_sheet = workbook['Chart 1a']
    hardchart1b_sheet = workbook['Chart 1b']
    hardchart1c_sheet = workbook['Chart 1c']
    hardchart2a_sheet = workbook['Chart 2a']
    hardchart2b_sheet = workbook['Chart 2b']
    hardchart2c_sheet = workbook['Chart 2c']
    charts_sheet = workbook['Chart 3 and 4']
    #all_charts_sheet = workbook['ALL Charts']

    #Add in headers and clean up the categorized sheet.
    cat_sheet.delete_cols(2)
    cat_sheet.insert_rows(1)
    cat_sheet["A1"] = "URL"
    cat_sheet["B1"] = "SEGMENT"
    cat_sheet["C1"] = "KEYWORDS"
    cat_sheet["D1"] = "SCORE"
    cat_sheet["E1"] = "SEGMENT"
    cat_sheet["F1"] = "KEYWORDS"
    cat_sheet["G1"] = "SCORE"
    cat_sheet["H1"] = "SEGMENT"
    cat_sheet["I1"] = "KEYWORDS"
    cat_sheet["J1"] = "SCORE"
    #Add in headers for the sections sheet. 
    sections_sheet['A1'] = "URL SECTION ONE"
    sections_sheet['B1'] = "URL SECTION TWO"
    #Add in headers for the hard chart 1 sheet
    hardchart1a_sheet['A1'] = "Most Common Section Out of All URLs"
    hardchart1a_sheet['A3'] = "Segments"
    hardchart1a_sheet['B3'] = "Count"
    hardchart2a_sheet['A1'] = "Most Common Segment Out of All URLs"
    hardchart2a_sheet['A3'] = "Sections"
    hardchart2a_sheet['B3'] = "Count"
    hardchart1b_sheet['A1'] = "Second Most Common Section Out of All URLs"
    hardchart1b_sheet['A3'] = "Segments"
    hardchart1b_sheet['B3'] = "Count"
    hardchart2b_sheet['A1'] = "Second Most Common Segment Out of All URLs"
    hardchart2b_sheet['A3'] = "Sections"
    hardchart2b_sheet['B3'] = "Count"
    hardchart1c_sheet['A1'] = "Third Most Common Section Out of All URLs"
    hardchart1c_sheet['A3'] = "Segments"
    hardchart1c_sheet['B3'] = "Count"
    hardchart2c_sheet['A1'] = "Third Most Common Segment Out of All URLs"
    hardchart2c_sheet['A3'] = "Sections"
    hardchart2c_sheet['B3'] = "Count"

    (http_true) = url_path_parser()
    (null_total,unsafe_total, safe_total, safe_segment_list, section_list) = url_totals()
    (most_popular_section, second_most_popular_section, third_most_popular_section) = section_many_segments(section_list)
    (most_popular_segment, second_most_popular_segment, third_most_popular_segment) = segment_many_sections(safe_segment_list, http_true)
    section_bar_chart(most_popular_section, second_most_popular_section, third_most_popular_section)
    segment_bar_chart(most_popular_segment, second_most_popular_segment, third_most_popular_segment)
    safe_pie_chart(null_total,unsafe_total, safe_total)
    popular_bar_chart(safe_segment_list)
    workbook.save(filename="{}{}{}".format("charts_",shorter_name,".xlsx"))
    print("Using input file named {} the context script has finished!".format(name))