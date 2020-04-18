#!/usr/bin/env python3

"""
cat.py is used to create four charts
from the url data in the csv files created from the internal tool 
called datashot (ie. grapeshot's url categorization tool). 
Each row in the csv file represents a URL with it's corresponding grapeshot
URL categorization data.
Author: Taylor Higgins
Last modified: 2/28/2020
"""
import sys
import re
import csv
import pandas as pd
import openpyxl
import operator
#import collections
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint


def url_path_parser():
    """Loop to parse all the URL path sections from the existing URLs, this is mostly needed for chart 3. And determine if http or www. """
    
    #add in try if so no error issue from josh TypeError
    for row in cat_sheet.iter_rows(min_col=1, min_row=2, max_col=1, max_row=cat_sheet.max_row):    
        for cell in row:
            #if column1 url in sheet 1
            #contains http://
            #do this logic for http:// url's
            #this below works for https
            #split_cell_list = cell.value.split('/')[3:]
            #else:
            #do this logic for www url's
            #print(type(cell.value)) #string
            #this below works for www.
            #print(type(cell.value))
            try:
                if 'http' in cell.value:
                    #print(type(cell.value))
                    #print(cell.value)
                    split_cell_list = cell.value.split('/')[3:]
                    http_true = 1
                else:
                    split_cell_list = cell.value.split('/')[1:]
                    http_true = 0
                tuple_section = (split_cell_list,)
            except TypeError:
                pass
            #print(http_true)
            #print(type(tuple_section)) #tuple
            for row in tuple_section:
                #print(type(row)) #list
                #print(row)
                sections_sheet.append(row)
                #data_sheet.append(row)
    return(http_true)

def standard_safety():
    """Function to show keywords and score for unsafe (if any row has a gv segment with score 15+) (and safe down the line) urls
    segments have their own nested dctionary under each url."""
    
    #convert tuple rows into list
    #list(rows)
    gv_unsafe_threshold = 15.00
    sheet_dictionary = {}
    total_row_count = cat_sheet.max_row
    i=0    
    for row in cat_sheet.iter_rows(min_col=1, min_row=2, max_col=40, max_row=cat_sheet.max_row): #40
        sheet_dictionary[i] = {}
        for cell in row:
            try:
                sheet_dictionary[i]["url"] = row[0].value
                sheet_dictionary[i]["segments"] = [row[1].value,row[4].value,row[7].value,row[10].value,row[13].value,row[16].value,row[19].value,row[22].value,row[25].value,row[28].value,row[31].value,row[34].value]
                sheet_dictionary[i]["keywords"] = [row[2].value, row[5].value, row[8].value,row[11].value, row[14].value, row[17].value,row[20].value, row[23].value, row[26].value,row[29].value, row[32].value, row[35].value]
                sheet_dictionary[i]["score"] = [row[3].value, row[6].value, row[9].value,row[12].value, row[15].value, row[18].value,row[21].value, row[24].value, row[27].value,row[30].value, row[33].value, row[36].value]
                sheet_dictionary[i]["safety_verdict"] = "unknown"
                sheet_dictionary[i]["safety_boolean"] = None
            except AttributeError:
                continue
        i+=1


    #TODO work on figuring out urls/rows that aren't safe, but don't have gv, null
    #go to file and see how many scores are more than 15.00
    #it's because they are duplicating, multipel gs is getting added twice.
    for key in sheet_dictionary.keys():
        #print(key)
        #print(sheet_dictionary[key]["url"])
        for segment in sheet_dictionary[key]["segments"]:
        #if gv is above threshhold entire row unsafe, break
        #if gv is below threshold, safe or undetermined, continue
        #if gs is present and gv isn't above threshold safe
        #if it's other than disregard that segment
        #we could disregard all rows with only garbage segments
        #then do the gv threshold test, and minus that total from url total to find safe
            try:
                if 'gv_' in segment:
                    #print(segment)
                    the_gv_index = sheet_dictionary[key]["segments"].index(segment)
                    #print(the_gv_index)
                    gv_score = float(sheet_dictionary[key]["score"][the_gv_index])
                    if gv_score >= gv_unsafe_threshold:
                        # print(type(gv_score))
                        # print(gv_score)
                        # print("unsafe")
                        sheet_dictionary[key]["safety_verdict"] = "unsafe gv score is {}".format(gv_score)
                        #HOORAY we just need get to adjust the total safe unsafe URL count using this logic, 
                        #mark this entire row as unsafe if there's gv > 15.00
                        sheet_dictionary[key]["safety_boolean"] = False
                        print(sheet_dictionary[key]["url"])
                        print(sheet_dictionary[key])
                        print(sheet_dictionary[key]["safety_verdict"])
                        print(sheet_dictionary[key]["safety_boolean"])
                        #TODO to get it to leave this row/this dict USE BREAK??
                        break
                    else:
                        # print(gv_score)
                        sheet_dictionary[key]["safety_verdict"] = "safe gv score is {}".format(gv_score)
                        sheet_dictionary[key]["safety_boolean"] = True
                        print(sheet_dictionary[key]["url"])
                        print(sheet_dictionary[key])
                        print(sheet_dictionary[key]["safety_verdict"])
                        print(sheet_dictionary[key]["safety_boolean"])
                        # print(gv_score)
                        # print("safe")
                elif 'gs_' in segment:
                    sheet_dictionary[key]["safety_verdict"] = "safe no gv segments in the url"
                    sheet_dictionary[key]["safety_boolean"] = True
                    print(sheet_dictionary[key]["url"])
                    print(sheet_dictionary[key])
                    print(sheet_dictionary[key]["safety_verdict"])
                    print(sheet_dictionary[key]["safety_boolean"])
                else:
                    sheet_dictionary[key]["safety_verdict"] = "null"
                    print(sheet_dictionary[key]["url"])
                    print(sheet_dictionary[key])
                    print(sheet_dictionary[key]["safety_verdict"])
                    print(sheet_dictionary[key]["safety_boolean"])
                    #print("Null")
            #TODO investigate this try except with skewing safe unsafe numbers
            except TypeError:
                continue
    #here or elsewhere perhaps up in safe null etc, we want to 
    #loop through all verdicts for all url dicts or rows
    #if it starts with safe count as safe, if it starts as unsafe count as unsafe
    
    #tests     
    #print(sheet_dictionary[26]) #should be safe but has a gv seg good
    #print(sheet_dictionary[9]) #should be unsafe good 
    #print(sheet_dictionary[21]) #will be safe
    return(sheet_dictionary)

def count_keywords_segments(sheet_dictionary):
    """TODO in its own tab for each i'll write the row of each dict that holds the total count"""
    #go through each row/url dict in the sheet_dict
    #if the row dict's key verdict value starts with safe
    #add all segments to safe segment count dict
    #and add all keywords after stripping of | to safe segment
    #if row's dict key verdict value starts with unsafe do the opposite
    #maybe add a step where we  
    #go through each segments list and see if starts with gv or gs
    #only add gs to safe and gv to unsafe 
    #use index to access segment and appropriate keywords

    #tests
    #print(sheet_dictionary[5])
    #print(type(sheet_dictionary[5]["safety_boolean"]))
    #unsafe_segment_count_sheet.append()

def url_totals(sheet_dictionary):

    """Find Totals for the charts. Categorized URLs total comes from the total row count in the categorized sheet 
    or by going through all the non gv_ and gx_ segments in the all sheet. Unsafe URLs Total
    comes from any url starting with gv_ we are updating it so that it would need a score
    over 15 via the safe verdict key from the sheet dictionary. TODO This feels fairly finished, need to test the new
    way of testing on some more files and put it in the graph. """             

    unsafe_total = 0
    null_total = 0
    actual_safe_total = 0
    fake_total = 0
    safe_segment_dict = {}
    section_dict = {}
    new_unsafe_total = 0
    new_safe_total = 0
    new_null_total = 0

    #TODO figure out if using IS here is skewing results on total safe unsafe
    #Work on new unsafe total and safe total here
    #loop through sheet dictionary, if url dict contains safety_verdict
    #starting with unsafe or safe add a point to the total
    for key in sheet_dictionary.keys():
        if sheet_dictionary[key]["safety_boolean"] is False:
            #print(sheet_dictionary[key]["safety_boolean"])
            #print("add to new_sheetdict_safe_total")
            #print(sheet_dictionary[key]["url"])
            new_unsafe_total +=1
        elif sheet_dictionary[key]["safety_boolean"] is True:
            #print(sheet_dictionary[key]["safety_boolean"])
            #print("add to new_sheetdict_unsafe_total")
            #print(sheet_dictionary[key]["url"])
            new_safe_total +=1
        elif sheet_dictionary[key]["safety_boolean"] is None:
            new_null_total+=1

    print("new unsafe total is {}".format(new_unsafe_total))
    print("new safe total is {}".format(new_safe_total))
    print("new null total is {}".format(new_null_total))


    #Starting loop to find uniquely null URL rows.
    for row in cat_sheet.iter_rows(min_col=2, min_row=2, max_col=cat_sheet.max_column, max_row=cat_sheet.max_row): 
        for cell in row:
            null_segment = re.compile(r'^gx_')
            try:
                check_null = null_segment.search(cell.value)
                if check_null is not None: 
                    null_total +=1 
                    #We remove break because we want any gx to be returned. 
                else:
                    continue

            except TypeError:
                #This should pass from the blank cell to the next in the row, then finally to next URL row.
                pass 

    #Starting loop to find uniquely unsafe URL rows.
    for row in cat_sheet.iter_rows(min_col=2, min_row=2, max_col=cat_sheet.max_column, max_row=cat_sheet.max_row): 
        for cell in row:
            #write this out so we ignore gv_safe
            fake_unsafe_segment = re.compile(r'^gv_safe')
            unsafe_segment = re.compile(r'^gv_')
            try:
                check_fake = fake_unsafe_segment.search(cell.value)
                check_safety = unsafe_segment.search(cell.value)
                if check_fake is not None:
                    #print(cell.value)
                    #print(row) #to check if safety is accurate
                    fake_total += 1
                    break
                elif check_safety is not None: 
                    #print(cell)
                    #keyword_of_segment = (cell + 1).value #if this could work somehow...
                    #print(cell + 1) # or <Cell 'Sheet1'.H3984> edit the string so it's <Cell 'Sheet1'.H3985>
                    #print(cell.value)
                    #print(row[0]) #to check if safety is accurate
                    unsafe_total +=1 
                    #We put a break here so that we don't over count URLs that were categorized multiple times as unsafe.
                    break 

            except TypeError:
                #This should pass from the blank cell to the next in the row, then finally to next URL row.
                pass 


    #Starting loop to create a dict of absolute count of segment appearances regardless of unique URL row. 
    for row in cat_sheet.iter_rows(min_col=2, min_row=2, max_col=cat_sheet.max_column, max_row=cat_sheet.max_row):
        for cell in row:
            safe_segment = re.compile(r'^gs_')
            try:
                check_safety = safe_segment.search(cell.value)
                if check_safety is not None: 
                    new_segment = cell.value
                    actual_safe_total += 1
                    #print(row) #to check if safety is accurate
                    if new_segment in safe_segment_dict: 
                        #If the segment is in the dict already add one to the current value.
                        safe_segment_dict[new_segment] = safe_segment_dict[new_segment] + 1
                    else:
                        #If the segment is not in the dict already then add it and make the value 1.
                        safe_segment_dict[new_segment] = 1
                    break #adding this break to see if we can get an accurate gs_ count
            except TypeError:
                pass

    #Here we are sorting the dict of safe segments in descending order by the value of the dict.   
    sorted_safe_seg_dict = dict(sorted(safe_segment_dict.items(), key=operator.itemgetter(1), reverse=True))
    #print(sorted_safe_seg_dict)
    #Here we are turning the sorted dict into a list so we can later append it by row and create the segment bar chart.
    safe_segment_list = list(sorted_safe_seg_dict.items())
    #print(safe_segment_list[0][0])

    #Starting loop to create a dict of absolute count of section appearances regardless of unique URL row or safety.
    for row in sections_sheet.iter_rows(min_col=1, min_row=2, max_col=sections_sheet.max_column, max_row=sections_sheet.max_row):
        #print(row)
        for cell in row:
            #print(cell.value)
            try:
                if cell.value in section_dict:
                    #If section is in the dict already add one to the current value. 
                    section_dict[cell.value] = section_dict[cell.value] + 1
                    #print(cell.value)
                else:
                    #If the segment is not in the dict already then add it and make the value 1.
                    section_dict[cell.value] = 1
            except TypeError:
                pass

    #Here we are sorting the dict of sections in descending order by the value of the dict.
    sorted_section_dict = dict(sorted(section_dict.items(), key=operator.itemgetter(1), reverse=True))

    #print(sorted_section_dict)
    #Here we are turning the sorted dict into a list so we can later append it by row to the file. 
    section_list = list(sorted_section_dict.items())
    
    
    #Calculate total categorized URLs. 
    row_count = cat_sheet.max_row 
    total_cat_urls = row_count - 1 - null_total
    safe_total = total_cat_urls - unsafe_total

    #print(row_count, actual_safe_total, unsafe_total, null_total, fake_total)

    charts_sheet['A1'] = 'Total URLs Categorized' 
    charts_sheet['B1'] = total_cat_urls
    # print("null total is {}".format(null_total))
    # print("unsafe total is {}".format(unsafe_total))
    # print("safe total is {}".format(safe_total))

    return(null_total, unsafe_total, safe_total, safe_segment_list, section_list)

def section_many_segments(section_list):
    """Create the Incremental Inventory Bar Chart"""
    #Here we pick the most popular section from the section list. 
    #Note that we might end up picking the top 3 to give more options.
    
    most_popular_section = section_list[2][0] #do the top 3
    #print(section_list[2][0])
    hardchart1_sheet['B1'] = most_popular_section
    popular_segment_count = {}

    #Here we loop through all rows in column 1 or URLs in cat_sheet
    #Currently this will loop through everything so there could be an issue.
    #If cell.value in a non column 1 col contains most_popular_section
    for row in cat_sheet.iter_rows(min_row=2,min_col=1,max_col=cat_sheet.max_column, max_row=cat_sheet.max_row):
        #For any cell in column 1 that contains the most popular section
        for cell in row:
            try:
                if most_popular_section in cell.value:
                    #for every 3 column after column 2 add the value of the cell to a dict and start counting it 
                    column_list = [2,5,8,11,14,17,20,23,26,29,32,35,38,41,44]
                    #dynamically loop through all segment columns
                    for column_number in column_list:
                        popular_segment = cat_sheet.cell(row=cell.row, column=column_number).value
                        #go through all cells in that row and pull out all gs segments
                        safe_segment = re.compile(r'^gs_')
                        try:
                            #check if that segment is in the dict yet 
                            check_seg_safety = safe_segment.search(popular_segment)
                            if check_seg_safety is not None:
                                if popular_segment in popular_segment_count:
                                    popular_segment_count[popular_segment] = popular_segment_count[popular_segment] + 1
                                else:
                                    popular_segment_count[popular_segment] = 1
                        except TypeError:
                            pass
                else:
                    continue
            except TypeError:
                pass

    popular_section_key = [most_popular_section]

    
    #this goes into a nested dict. {1_section:{1seg:1segcount,2seg:2segcount,3seg:3segcount}}
    nested_popular_segment_count = dict.fromkeys(popular_section_key,popular_segment_count)

    #find percentage of each key or top 5 most popular keys in nested_popular_segment_count
    #Here we are appending the section list to the section tab. Moved from url totals function. 
    #for row in section_list:
        #sections_sheet.append(row)
        #trying to label that
        #data_sheet[data_sheet.max_row]
        #data_sheet.append(row)

    sorted_popular_segment_count = dict(sorted(popular_segment_count.items(), key=operator.itemgetter(1), reverse=True))
    #making a list popular_segment_count
    sorted_popular_seg_list = list(sorted_popular_segment_count.items())

    #appending sorted_popular_seg_list to sections sheet
    for row in sorted_popular_seg_list:
        #sections_sheet.append(row)
        #data_sheet.append(row)
        hardchart1_sheet.append(row)

    bar = BarChart()
    labels = Reference(hardchart1_sheet, min_col=1, min_row=4, max_row=14)
    data = Reference(worksheet=hardchart1_sheet, min_col=2, min_row=4, max_row=14)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = 'the "{}" section has many segments'.format(most_popular_section) 

    hardchart1_sheet.add_chart(bar, "E3")

def segment_many_sections(safe_segment_list, http_true):
    """Create SubDomain Bar Chart"""
    #print(safe_segment_list)
    most_popular_segment = safe_segment_list[0][0]
    hardchart2_sheet['B1'] = most_popular_segment
    popular_section_count = {}

    #Here we loop through all rows/segment columns in cat sheets to find any rows with the popular segment 
    for row in cat_sheet.iter_rows(min_row=2,min_col=1,max_col=cat_sheet.max_column, max_row=cat_sheet.max_row):
        for cell in row:
            try:
                if most_popular_segment in cell.value:
                #When we find a row with the popular segment
                #We need to find what section it is in
                #by going to first column of that row
                #or row number
                #finding section info from parser or url and parse again
                    #print(row[0].value)
                    url_to_be_parsed_again = row[0].value
                    #print(url_to_be_parsed_again)
                    #if column1 url in sheet 1
                    #contains http://
                    #do this logic for http:// url's
                    try:
                        if http_true == 1:
                            popular_section_entire = url_to_be_parsed_again.split('/')[3:]
                            #print(cell.value)
                            #print(url_to_be_parsed_again)
                            #print(popular_section_entire) #check if there is something in first IndexError

                            popular_section = popular_section_entire[0]
                            #print("hi we're in the if")

                            if popular_section in popular_section_count:
                                popular_section_count[popular_section] = popular_section_count[popular_section] + 1
                                #print("hi we're in the nested if")
                            else:
                                popular_section_count[popular_section] = 1
                                #print("hi we're in the nested else")
                        #else:
                        #do this logic for www url's
                        else:
                            popular_section_entire = url_to_be_parsed_again.split('/')
                            popular_section = popular_section_entire[1]
                            #print("hi we're in the else")

                            if popular_section in popular_section_count:
                                popular_section_count[popular_section] = popular_section_count[popular_section] + 1
                                #print("hi we're in the nested if")
                            else:
                                popular_section_count[popular_section] = 1
                                #print("hi we're in the nested else")
                    except IndexError:
                        pass
                else:       
                    continue
            except TypeError:
                pass
    #print(most_popular_segment)
    #print(popular_section_count)

    #then we can go through similarly as we did in the first chart 
    sorted_popular_section_count = dict(sorted(popular_section_count.items(), key=operator.itemgetter(1), reverse=True))
    #making a list popular_section_count
    sorted_popular_section_count = list(sorted_popular_section_count.items())
    #print(type(sorted_popular_section_count))
    #print((sorted_popular_section_count))
    for row in sorted_popular_section_count:
        hardchart2_sheet.append(row)

    #here is a bar chart for chart 2
    bar = BarChart()
    labels = Reference(hardchart2_sheet, min_col=1, min_row=4, max_row=14)
    data = Reference(worksheet=hardchart2_sheet, min_col=2, min_row=4, max_row=14)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = 'sections where the "{}" segment is'.format(most_popular_segment) 
    hardchart2_sheet.add_chart(bar, "E3")

def safe_pie_chart(null_total, unsafe_total, safe_total):
    """Create Safe Unsafe Pie Chart"""

    #Here we create the data table and appending it in the Charts Sheet.
    #TODO: make this a percent! 
    safe_unsafe_data = [['Safe', safe_total], ['Unsafe', unsafe_total]]
    for row in safe_unsafe_data:
        charts_sheet.append(row)

    #Here we set up the pie chart using openpyxl. 
    pie = PieChart()
    labels = Reference(charts_sheet, min_col=1, min_row=2, max_row=3)
    data = Reference(charts_sheet, min_col=2, min_row=1, max_row=3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Unsafe URLs"
    charts_sheet.add_chart(pie, "H2")

def popular_bar_chart(safe_segment_list):
    """Create Popular Segments Bar Chart"""

    charts_sheet['A5'] = 'Segment'
    charts_sheet['B5'] = 'Total Appearances'

    #Here we append the total segment data to the Charts Sheet.
    for row in safe_segment_list:
        charts_sheet.append(row)

    #Here we set up a bar chart using openpyxl.
    bar = BarChart()
    labels = Reference(charts_sheet, min_col=1, min_row=6, max_row=21)
    data = Reference(worksheet=charts_sheet, min_col=2, min_row=6, max_row=21)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = 'Total Appearances'
    charts_sheet.add_chart(bar, "H20")


if __name__ == '__main__': 

    print('Hello! The context script is working, one moment please...')
    #Open csv file from Datashot using the command line argument.
    name = sys.argv[1] #meredith.csv.cat.csv
    cleaned_name = name.split(".")
    shorter_name = cleaned_name[0] 
    #Transform the csv from datashot to xlsx
    #issue with unshortened so added sep and header, could maybe add: error_bad_lines=False
    #added engine='python' from error suggestion:  ParserWarning: Falling back to the 'python' engine because the 'c' engine does not support regex separator
    read_file = pd.read_csv(name, delimiter=',', header=None, engine='python', names=list(range(500))) 
    read_file.to_excel("{}{}{}".format("charts_",shorter_name,".xlsx"),index=None, header=False)
    filename = "{}{}{}".format("charts_",shorter_name,".xlsx")
    workbook = load_workbook(filename) 

    workbook.create_sheet('URL Sections')
    workbook.create_sheet('Chart 1')
    workbook.create_sheet('Chart 2')
    workbook.create_sheet('Chart 3 and 4')
    workbook.create_sheet('Unsafe Segment Count')
    workbook.create_sheet('Unsafe Keywords Count')
    workbook.create_sheet('Safe Segment Count')
    workbook.create_sheet('Safe Keywords Count')

    #Create new sheets for sections and charts and make variables of each sheet we'll need. 
    cat_sheet = workbook['Sheet1']
    sections_sheet = workbook['URL Sections']
    hardchart1_sheet = workbook['Chart 1']
    hardchart2_sheet = workbook['Chart 2']
    charts_sheet = workbook['Chart 3 and 4']
    unsafe_segment_count_sheet = workbook['Unsafe Segment Count']
    unsafe_keywords_count_sheet = workbook['Unsafe Keywords Count']
    safe_segment_count_sheet = workbook['Safe Segment Count']
    safe_keywords_count_sheet = workbook['Safe Keywords Count']

    #Add in headers and clean up the categorized sheet.
    cat_sheet.delete_cols(2)
    cat_sheet.insert_rows(1)
    cat_sheet["A1"] = "URL"
    cat_sheet["B1"] = "SEGMENT"
    cat_sheet["C1"] = "KEYWORDS"
    cat_sheet["D1"] = "SCORE"
    cat_sheet["E1"] = "SEGMENT"
    cat_sheet["F1"] = "KEYWORDS"
    cat_sheet["G1"] = "SCORE"
    cat_sheet["H1"] = "SEGMENT"
    cat_sheet["I1"] = "KEYWORDS"
    cat_sheet["J1"] = "SCORE"
    #Add in headers for the sections sheet. 
    sections_sheet['A1'] = "URL SECTION ONE"
    sections_sheet['B1'] = "URL SECTION TWO"
    #Add in headers for the hard chart 1 sheet
    hardchart1_sheet['A1'] = "Most Common Section Out of All URLs"
    hardchart1_sheet['A3'] = "Segments"
    hardchart1_sheet['B3'] = "Count"
    hardchart2_sheet['A1'] = "Most Common Segment Out of All URLs"
    hardchart2_sheet['A3'] = "Sections"
    hardchart2_sheet['B3'] = "Count"
    unsafe_segment_count_sheet['A1'] = "Segment"
    unsafe_segment_count_sheet['A2'] = "Count"
    unsafe_keywords_count_sheet['A1'] = "Keywords"
    unsafe_keywords_count_sheet['A2'] = "Count"
    safe_segment_count_sheet['A1'] = "Segment"
    safe_segment_count_sheet['A2'] = "Count"
    safe_keywords_count_sheet['A1'] = "Keywords"
    safe_keywords_count_sheet['A2'] = "Count"

    (http_true) = url_path_parser()
    (sheet_dictionary) = standard_safety()
    (null_total,unsafe_total, safe_total, safe_segment_list, section_list) = url_totals(sheet_dictionary)
    count_keywords_segments(sheet_dictionary)
    segment_many_sections(safe_segment_list, http_true)
    section_many_segments(section_list)
    safe_pie_chart(null_total,unsafe_total, safe_total)
    popular_bar_chart(safe_segment_list)
    workbook.save(filename="{}{}{}".format("charts_",shorter_name,".xlsx"))
    print("Using input file named {} the context script has finished!".format(name))